import openpyxl
import datetime

column = ['N', 'O', 'F', 'G', 'I', 'H']
name_key = ['start', 'end', 'mileage', 'duration', 'between', 'halt']
format_dat = ['%Y-%m-%d', '%d.%m %Y', '%d %m.%Y', '%d %m %Y', ' %d.%m.%Y']

def the_end(list_for_fillings, file_name):
    car_missing = []
    book = openpyxl.open(file_name)
    sheet = book.active
    cell_number = sheet['C8':f'C{sheet.max_row}']
    cell_date = sheet['B8':f'B{sheet.max_row}']
    dictionary_car_missing = {
        'nm': '',
        'date': '',
    }

    def search_coordinate(car_number):
        coordin = ''
        for tuple in cell_number:
            for value in tuple:
                if value.value == car_number:
                    coordin = value.coordinate
                    return coordin

    def data_format(data):
        data = str(data)
        data = data.replace(' 00:00:00', '')
        print(data)
        for format in format_dat:
            try:
                data = datetime.datetime.strptime(data, format).date()
                return data
            except ValueError:
                continue
    # def completion(column, name_key, row):
    #     print(str(sheet[f'B{row}'].value)[:10])
    #     print(data['date'])
    #     # if str(data['date']) == str(sheet[f'B{row}'].value)[:10]:
    #     sheet[f'{column}{row}'] = date[name_key]

    def search(nomer, date):
        for tuple_1, tuple_2 in zip(cell_number, cell_date):
            for value_1, value_2 in zip(tuple_1, tuple_2):
                # print(value_1.value, value_2.value, nomer, date)
                value_2.value = data_format(value_2.value)
                if value_1.value == nomer and str(value_2.value)\
                        == date:
                    # print(value_1.row)
                    return value_1.row


    # for car in list_for_fillings:
    #     car_number = car[0]['nm']
    #     coordin = search_coordinate(car_number)
    #     coordin = sheet[f'{coordin}'].row
    #     # for data in car:
    #     #     if data != {}:
    #     #         for letter, key in zip(column, name_key):
    #     #             completion(letter, key, coordin)
    #     #     coordin += 1
    #     for date in car:
    #         if date == {}:
    #             break
    #         for data in sheet[f'B{coordin}':f'B{sheet.max_row}']:
    #             print(sheet[f'B{coordin}'].value)
    #             if sheet[f'B{coordin}'].value == None:
    #                 break
    #             if sheet[f'B{coordin}'].value == date['date']:
    #                 for letter, key in zip(column, name_key):
    #                     completion(letter, key, coordin)

    for car in list_for_fillings:
        for date in car:
            print(date)
            if date == {}:
                break
            coordin = search(date['nm'], date['date'])
            for letter, key in zip(column, name_key):
                sheet[f'{letter}{coordin}'] = date[key]

    for tuple_1, tuple_2 in zip(cell_number, cell_date):
        for value_1, value_2 in zip(tuple_1, tuple_2):
            if value_1.value != None:
                date_coordin = value_1.coordinate
                date_coordin = sheet[date_coordin].offset(row=0, column=3).coordinate
                if sheet[date_coordin].value == None:
                    dictionary_car_missing['nm'] = value_1.value
                    dictionary_car_missing['date'] = str(value_2.value)[:10]
                    car_missing.append(dictionary_car_missing.copy())


    book.save(file_name)
    book.close()
    return car_missing


